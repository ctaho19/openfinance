#!/usr/bin/env python3
"""
Chris Taho - Financial Freedom Model 2025-2026
Every dollar has a job. Debt avalanche strategy.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
import math

wb = Workbook()

# Color constants (finance industry standard)
BLUE_FONT = Font(color="0000FF")  # Inputs/assumptions
BLACK_FONT = Font(color="000000", bold=False)  # Formulas
GREEN_FONT = Font(color="008000")  # Cross-sheet links
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
ALERT_FILL = PatternFill("solid", fgColor="FFCCCC")
SUCCESS_FILL = PatternFill("solid", fgColor="CCFFCC")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_currency(ws, cell, is_input=False):
    ws[cell].number_format = '"$"#,##0.00'
    if is_input:
        ws[cell].font = BLUE_FONT
        ws[cell].fill = INPUT_FILL

def format_percent(ws, cell, is_input=False):
    ws[cell].number_format = '0.00%'
    if is_input:
        ws[cell].font = BLUE_FONT
        ws[cell].fill = INPUT_FILL

def set_header(ws, cell, value):
    ws[cell] = value
    ws[cell].font = HEADER_FONT
    ws[cell].fill = HEADER_FILL
    ws[cell].alignment = Alignment(horizontal='center')

# ============== SHEET 1: SNAPSHOT ==============
ws1 = wb.active
ws1.title = "Snapshot"
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 50

# Title
ws1['A1'] = "FINANCIAL SNAPSHOT - DECEMBER 27, 2025"
ws1['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws1.merge_cells('A1:D1')

# Current Reality
ws1['A3'] = "CURRENT REALITY"
ws1['A3'].font = Font(bold=True, size=12)

ws1['A4'] = "Cash on Hand (as of today)"
ws1['B4'] = 250
format_currency(ws1, 'B4', True)

ws1['A5'] = "Net Paycheck (biweekly)"
ws1['B5'] = 3049
format_currency(ws1, 'B5', True)

ws1['A6'] = "Monthly Income (approx)"
ws1['B6'] = "=B5*26/12"
format_currency(ws1, 'B6')

ws1['A7'] = "Annual Income"
ws1['B7'] = "=B5*26"
format_currency(ws1, 'B7')

# Total Debt by Category
ws1['A9'] = "TOTAL DEBT BREAKDOWN"
ws1['A9'].font = Font(bold=True, size=12)

set_header(ws1, 'A10', "Debt Category")
set_header(ws1, 'B10', "Balance")
set_header(ws1, 'C10', "Min Monthly")
set_header(ws1, 'D10', "Notes")

debts = [
    ("Credit Cards (HIGH PRIORITY)", 15840.59, 659.98, "24-30% APR - Pay these FIRST"),
    ("Auto Loans", 20364.40, 638.90, "BMW is PAST DUE - $1,204 behind"),
    ("Student Loans", 39666.03, 528.93, "Nelnet deferred until Feb 2026"),
    ("BNPL (Buy Now Pay Later)", 6237.85, 2154.84, "Self-liquidating by ~Oct 2026"),
]

row = 11
for name, balance, minpay, notes in debts:
    ws1[f'A{row}'] = name
    ws1[f'B{row}'] = balance
    format_currency(ws1, f'B{row}')
    ws1[f'C{row}'] = minpay
    format_currency(ws1, f'C{row}')
    ws1[f'D{row}'] = notes
    if "PAST DUE" in notes or "HIGH" in notes:
        ws1[f'A{row}'].fill = ALERT_FILL
    row += 1

ws1[f'A{row}'] = "TOTAL DEBT"
ws1[f'A{row}'].font = Font(bold=True)
ws1[f'B{row}'] = f"=SUM(B11:B{row-1})"
ws1[f'B{row}'].font = Font(bold=True)
format_currency(ws1, f'B{row}')
ws1[f'C{row}'] = f"=SUM(C11:C{row-1})"
ws1[f'C{row}'].font = Font(bold=True)
format_currency(ws1, f'C{row}')
total_debt_row = row

# Monthly Budget Summary
row += 2
ws1[f'A{row}'] = "MONTHLY CASH FLOW"
ws1[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws1[f'A{row}'] = "Monthly Income"
ws1[f'B{row}'] = "=B6"
format_currency(ws1, f'B{row}')

row += 1
ws1[f'A{row}'] = "(-) Fixed Bills (Insurance, Subscriptions)"
ws1[f'B{row}'] = -287.97
format_currency(ws1, f'B{row}', True)

row += 1
ws1[f'A{row}'] = "(-) Debt Minimum Payments"
ws1[f'B{row}'] = f"=-C{total_debt_row}"
format_currency(ws1, f'B{row}')

row += 1
income_row = row - 3
bills_row = row - 2
debt_row = row - 1
ws1[f'A{row}'] = "= Available for Spending/Savings/Extra Debt"
ws1[f'B{row}'] = f"=B{income_row}+B{bills_row}+B{debt_row}"
ws1[f'B{row}'].font = Font(bold=True, color="008000")
format_currency(ws1, f'B{row}')
surplus_row = row

# Living Expenses Estimate
row += 2
ws1[f'A{row}'] = "ESTIMATED LIVING EXPENSES"
ws1[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws1[f'A{row}'] = "Food/Groceries"
ws1[f'B{row}'] = 400
format_currency(ws1, f'B{row}', True)

row += 1
ws1[f'A{row}'] = "Gas/Transportation"
ws1[f'B{row}'] = 200
format_currency(ws1, f'B{row}', True)

row += 1
ws1[f'A{row}'] = "Personal/Misc"
ws1[f'B{row}'] = 150
format_currency(ws1, f'B{row}', True)

row += 1
food_row = row - 3
gas_row = row - 2
misc_row = row - 1
ws1[f'A{row}'] = "Total Variable Spending"
ws1[f'B{row}'] = f"=SUM(B{food_row}:B{misc_row})"
format_currency(ws1, f'B{row}')
living_row = row

# TRUE remaining for extra debt paydown
row += 2
ws1[f'A{row}'] = "TRUE MONTHLY SURPLUS FOR DEBT ATTACK"
ws1[f'A{row}'].font = Font(bold=True)
ws1[f'B{row}'] = f"=B{surplus_row}-B{living_row}"
ws1[f'B{row}'].font = Font(bold=True, size=14)
ws1[f'B{row}'].fill = SUCCESS_FILL
format_currency(ws1, f'B{row}')


# ============== SHEET 2: DEBT DETAIL ==============
ws2 = wb.create_sheet("Debt Detail")
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 20
ws2.column_dimensions['G'].width = 15

ws2['A1'] = "COMPLETE DEBT INVENTORY - Sorted by Interest Rate (Avalanche Method)"
ws2['A1'].font = Font(bold=True, size=14, color="1F4E79")
ws2.merge_cells('A1:G1')

headers = ["Account Name", "Balance", "APR %", "Min Payment", "Due Day", "Status", "Priority"]
for col, h in enumerate(headers, 1):
    set_header(ws2, f"{get_column_letter(col)}3", h)

# All debts sorted by APR (avalanche method)
all_debts = [
    # Credit Cards (highest priority)
    ("Best Buy Credit Card", 6352.18, 29.99, 230, 8, "CURRENT", 1),
    ("Bank of America BankAmericacard", 5022.60, 24.99, 239.98, 8, "CURRENT", 2),
    ("American Express Green Card", 3978.32, 23.99, 170, 14, "CURRENT", 3),
    ("Navy Federal CashRewards", 487.49, 17.97, 20, 20, "CURRENT", 4),
    
    # BNPL with interest
    ("Affirm - 10 payments (due 28th)", 532.54, 32.43, 59.21, 28, "CURRENT", 5),
    ("Affirm - 4 payments left (due 28th)", 164.20, 32.53, 54.85, 28, "CURRENT", 6),
    ("Affirm - 3 payments left (due 29th)", 180.59, 32.40, 91.38, 29, "CURRENT", 7),
    ("Affirm - 10 payments (due 14th)", 591.83, 32.43, 59.21, 14, "CURRENT", 8),
    ("Affirm - Final Jan 21", 259.91, 22.52, 259.91, 21, "CURRENT", 9),
    ("Affirm - Final Jan 21 (B)", 34.60, 22.52, 34.60, 21, "CURRENT", 10),
    
    # Auto Loans
    ("Navy Federal Auto Loan", 2740.10, 16.85, 201.53, 20, "DEFERRED to Jul 2027", 11),
    ("LendKey Student Loan", 6214.58, 11.20, 110.69, 1, "PAST DUE - $111", 12),
    ("BMW Financial Services", 17624.30, 8.98, 437.37, 4, "PAST DUE - $1,204", 13),
    ("Sallie Mae", 11247.88, 8.13, 192.10, 7, "CURRENT", 14),
    ("Nelnet Federal Student Loans", 22203.57, 5.00, 226.14, 10, "DEFERRED to Feb 2026", 15),
    
    # 0% BNPL (pay minimums, they self-liquidate)
    ("Affirm - 12 payments (due 8th)", 947.28, 0, 78.94, 8, "CURRENT", 16),
    ("Affirm - 10 payments left (due 27th)", 710.22, 0, 78.95, 27, "CURRENT", 17),
    ("Affirm - 5 payments (due 17th)", 500, 0, 100, 17, "CURRENT", 18),
    ("Affirm - 4 payments left (due 27th)", 400.01, 0, 133.33, 27, "CURRENT", 19),
    ("Affirm - 12 payments (due 8th) B", 355.26, 0, 29.60, 8, "CURRENT", 20),
    ("Affirm - 12 payments (due 10th)", 355.25, 0, 29.60, 10, "CURRENT", 21),
    ("Affirm - 4 payments left", 250.01, 0, 83.33, 27, "CURRENT", 22),
    ("AfterPay - Misc Travel", 190.55, 0, 190.55, 19, "CURRENT", 23),
    ("Affirm - 3 payments left (due 24th)", 184.58, 0, 93.21, 24, "CURRENT", 24),
    ("Sezzle - Clearcover (2)", 179.25, 0, 59.75, 7, "CURRENT", 25),
    ("Sezzle - Jomashop", 126.42, 0, 63.21, 2, "CURRENT", 26),
    ("Sezzle - Expedia", 108.75, 0, 108.75, 10, "CURRENT", 27),
    ("Zip - Fashion Nova For Mom", 98.60, 0, 49.30, 29, "CURRENT", 28),
    ("Zip - Vent De Sud", 47.00, 0, 23.50, 2, "CURRENT", 29),
    ("Zip - Cash 3", 21.00, 0, 21.00, 18, "CURRENT", 30),
]

row = 4
for name, balance, apr, minpay, due, status, priority in all_debts:
    ws2[f'A{row}'] = name
    ws2[f'B{row}'] = balance
    format_currency(ws2, f'B{row}')
    ws2[f'C{row}'] = apr / 100
    format_percent(ws2, f'C{row}')
    ws2[f'D{row}'] = minpay
    format_currency(ws2, f'D{row}')
    ws2[f'E{row}'] = due
    ws2[f'F{row}'] = status
    ws2[f'G{row}'] = priority
    
    if "PAST DUE" in status:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws2[f'{col}{row}'].fill = ALERT_FILL
    elif priority <= 4:
        ws2[f'A{row}'].fill = PatternFill("solid", fgColor="FFE4B5")  # Orange for high priority
    row += 1

# Total
ws2[f'A{row}'] = "TOTAL"
ws2[f'A{row}'].font = Font(bold=True)
ws2[f'B{row}'] = f"=SUM(B4:B{row-1})"
format_currency(ws2, f'B{row}')
ws2[f'B{row}'].font = Font(bold=True)
ws2[f'D{row}'] = f"=SUM(D4:D{row-1})"
format_currency(ws2, f'D{row}')
ws2[f'D{row}'].font = Font(bold=True)


# ============== SHEET 3: EVERY-DOLLAR BUDGET ==============
ws3 = wb.create_sheet("Every Dollar Budget")
ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 45

ws3['A1'] = "BIWEEKLY PAY PERIOD BUDGET - Every Dollar Assigned"
ws3['A1'].font = Font(bold=True, size=14, color="1F4E79")
ws3.merge_cells('A1:D1')

ws3['A3'] = "This budget is for ONE pay period (2 weeks). You get paid biweekly."
ws3['A3'].font = Font(italic=True)

row = 5
ws3[f'A{row}'] = "INCOME"
ws3[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws3[f'A{row}'] = "Net Paycheck"
ws3[f'B{row}'] = 3049
format_currency(ws3, f'B{row}', True)
income_cell = row

row += 2
ws3[f'A{row}'] = "ALLOCATIONS (Bills + Debt Minimums)"
ws3[f'A{row}'].font = Font(bold=True, size=12)

# Calculate per-paycheck amounts (biweekly = 2.17 paychecks/month)
allocations = [
    ("BMW Auto Loan (CATCH UP FIRST!)", 437.37, "Due 4th - You're $1,204 behind"),
    ("Sallie Mae", 192.10, "Due 7th - Current, keep it that way"),
    ("LendKey (CATCH UP!)", 110.69, "Due 1st - You're $111 behind"),
    ("Best Buy Credit Card", 230, "Due 8th - PRIORITY: Highest APR"),
    ("Bank of America", 239.98, "Due 8th - PRIORITY: 2nd highest APR"),
    ("Amex Green Card", 170, "Due 14th"),
    ("Navy Federal Credit Card", 20, "Due 20th - Small balance, pay off soon"),
    ("Clearcover Insurance", 239 / 2.17, "Due 23rd - Divide by 2.17 for biweekly"),
    ("Subscriptions (iCloud, Prime, Google)", (15.99 + 9.99 + 2.99) / 2.17, "Combined subscriptions"),
    ("Gym", 20 / 2.17, "AMK Capital One Fitness"),
]

row += 1
set_header(ws3, f'A{row}', "Expense")
set_header(ws3, f'B{row}', "Per Paycheck")
set_header(ws3, f'C{row}', "Monthly Equiv")
set_header(ws3, f'D{row}', "Notes")

alloc_start = row + 1
for name, amount, notes in allocations:
    row += 1
    ws3[f'A{row}'] = name
    ws3[f'B{row}'] = amount
    format_currency(ws3, f'B{row}')
    ws3[f'C{row}'] = f"=B{row}*2.17"
    format_currency(ws3, f'C{row}')
    ws3[f'D{row}'] = notes
    if "CATCH UP" in name or "PRIORITY" in notes:
        ws3[f'A{row}'].fill = ALERT_FILL

alloc_end = row

# BNPL - These vary by paycheck but roughly:
row += 2
ws3[f'A{row}'] = "BNPL PAYMENTS (varies by period)"
ws3[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws3[f'A{row}'] = "Average BNPL payments per paycheck"
ws3[f'B{row}'] = 2154.84 / 2.17  # ~$993 per paycheck
format_currency(ws3, f'B{row}', True)
ws3[f'D{row}'] = "Check OpenFi for exact amounts each period"
bnpl_row = row

# Living Expenses
row += 2
ws3[f'A{row}'] = "LIVING EXPENSES"
ws3[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws3[f'A{row}'] = "Groceries/Food"
ws3[f'B{row}'] = 200
format_currency(ws3, f'B{row}', True)
food_row = row

row += 1
ws3[f'A{row}'] = "Gas"
ws3[f'B{row}'] = 100
format_currency(ws3, f'B{row}', True)
gas_row = row

row += 1
ws3[f'A{row}'] = "Personal/Misc"
ws3[f'B{row}'] = 75
format_currency(ws3, f'B{row}', True)
misc_row = row

# Summary
row += 2
ws3[f'A{row}'] = "SUMMARY"
ws3[f'A{row}'].font = Font(bold=True, size=12)

row += 1
ws3[f'A{row}'] = "Income"
ws3[f'B{row}'] = f"=B{income_cell}"
format_currency(ws3, f'B{row}')

row += 1
ws3[f'A{row}'] = "(-) Fixed Allocations"
ws3[f'B{row}'] = f"=-SUM(B{alloc_start}:B{alloc_end})"
format_currency(ws3, f'B{row}')

row += 1
ws3[f'A{row}'] = "(-) BNPL Payments"
ws3[f'B{row}'] = f"=-B{bnpl_row}"
format_currency(ws3, f'B{row}')

row += 1
ws3[f'A{row}'] = "(-) Living Expenses"
ws3[f'B{row}'] = f"=-SUM(B{food_row}:B{misc_row})"
format_currency(ws3, f'B{row}')

row += 1
summary_start = row - 4
ws3[f'A{row}'] = "= EXTRA FOR DEBT ATTACK OR SAVINGS"
ws3[f'A{row}'].font = Font(bold=True)
ws3[f'B{row}'] = f"=SUM(B{summary_start}:B{row-1})"
ws3[f'B{row}'].font = Font(bold=True, size=14)
ws3[f'B{row}'].fill = SUCCESS_FILL
format_currency(ws3, f'B{row}')


# ============== SHEET 4: 2026 PROJECTIONS ==============
ws4 = wb.create_sheet("2026 Projections")
ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 12

ws4['A1'] = "MONTH-BY-MONTH PROJECTION TO END OF 2026"
ws4['A1'].font = Font(bold=True, size=14, color="1F4E79")
ws4.merge_cells('A1:N1')

ws4['A2'] = "Assumptions: Minimum payments + $500/mo extra toward highest-APR debt (debt avalanche)"
ws4['A2'].font = Font(italic=True)

# Monthly projection headers
set_header(ws4, 'A4', "#")
set_header(ws4, 'B4', "Month")
set_header(ws4, 'C4', "Income")
set_header(ws4, 'D4', "Bills")
set_header(ws4, 'E4', "Debt Mins")
set_header(ws4, 'F4', "BNPL")
set_header(ws4, 'G4', "Living")
set_header(ws4, 'H4', "Extra Debt $")
set_header(ws4, 'I4', "Savings $")
set_header(ws4, 'J4', "Credit Cards")
set_header(ws4, 'K4', "Auto Loans")
set_header(ws4, 'L4', "Student Loans")
set_header(ws4, 'M4', "BNPL Bal")
set_header(ws4, 'N4', "Total Debt")
set_header(ws4, 'O4', "Emergency Fund")

for col in range(3, 16):
    ws4.column_dimensions[get_column_letter(col)].width = 13

# Model each month from Jan 2026 to Dec 2026
# Starting balances
cc_balance = 15840.59
auto_balance = 20364.40
student_balance = 39666.03
bnpl_balance = 6237.85
emergency_fund = 0

monthly_income = 6616.33
monthly_bills = 287.97
monthly_living = 750

# BNPL self-liquidating schedule (rough - decreases each month as loans end)
bnpl_schedule = [1800, 1600, 1500, 1300, 1100, 900, 700, 500, 300, 200, 100, 0]  # Approx monthly BNPL payments

row = 5
for month_num in range(1, 25):  # Jan 2026 to Dec 2027
    if month_num <= 12:
        month_name = datetime(2026, month_num, 1).strftime("%b %Y")
    else:
        month_name = datetime(2027, month_num - 12, 1).strftime("%b %Y")
    
    ws4[f'A{row}'] = month_num
    ws4[f'B{row}'] = month_name
    
    # Income
    ws4[f'C{row}'] = monthly_income
    format_currency(ws4, f'C{row}')
    
    # Fixed bills
    ws4[f'D{row}'] = monthly_bills
    format_currency(ws4, f'D{row}')
    
    # Debt minimums (excluding BNPL which is separate)
    # CC: 660, Auto: ~437 (BMW only after getting current), Student: ~530
    debt_mins = 660 + 437.37 + 302.79  # CC + BMW + active student loans
    if month_num >= 3:  # Nelnet comes out of deferment Feb 2026
        debt_mins += 226.14
    ws4[f'E{row}'] = debt_mins
    format_currency(ws4, f'E{row}')
    
    # BNPL (decreasing over time)
    bnpl_payment = bnpl_schedule[min(month_num - 1, 11)] if month_num <= 12 else 0
    ws4[f'F{row}'] = bnpl_payment
    format_currency(ws4, f'F{row}')
    
    # Living expenses
    ws4[f'G{row}'] = monthly_living
    format_currency(ws4, f'G{row}')
    
    # Calculate surplus
    if row == 5:
        surplus_formula = f"=C{row}-D{row}-E{row}-F{row}-G{row}"
    else:
        surplus_formula = f"=C{row}-D{row}-E{row}-F{row}-G{row}"
    
    # Split surplus: 80% debt, 20% savings until $1000 emergency fund
    ws4[f'H{row}'] = f"=MAX(0,(C{row}-D{row}-E{row}-F{row}-G{row})*0.8)"
    format_currency(ws4, f'H{row}')
    
    ws4[f'I{row}'] = f"=MAX(0,(C{row}-D{row}-E{row}-F{row}-G{row})*0.2)"
    format_currency(ws4, f'I{row}')
    
    # Debt balances (simplified projections with interest)
    if row == 5:
        # Starting balances
        ws4[f'J{row}'] = f"=15840.59-H{row}"  # CC
        ws4[f'K{row}'] = f"=20364.40-MAX(0,H{row}-J{row})"  # Auto (after CC paid)
        ws4[f'L{row}'] = 39666.03  # Student (pay minimums)
        ws4[f'M{row}'] = f"=6237.85-F{row}"  # BNPL
    else:
        # CC: Pay down with extra, add ~2% monthly interest on remaining
        ws4[f'J{row}'] = f"=MAX(0,J{row-1}*1.02-H{row})"
        # Auto stays same (paying minimums)
        ws4[f'K{row}'] = f"=K{row-1}*1.0075-437.37"  # ~9% APR
        # Student stays same (paying minimums)  
        ws4[f'L{row}'] = f"=L{row-1}*1.005-302.79"  # ~6% APR average
        # BNPL decreases by payment
        ws4[f'M{row}'] = f"=MAX(0,M{row-1}-F{row})"
    
    format_currency(ws4, f'J{row}')
    format_currency(ws4, f'K{row}')
    format_currency(ws4, f'L{row}')
    format_currency(ws4, f'M{row}')
    
    # Total debt
    ws4[f'N{row}'] = f"=J{row}+K{row}+L{row}+M{row}"
    format_currency(ws4, f'N{row}')
    
    # Emergency fund
    if row == 5:
        ws4[f'O{row}'] = f"=I{row}"
    else:
        ws4[f'O{row}'] = f"=MIN(1000,O{row-1}+I{row})"  # Cap at $1000 initial goal
    format_currency(ws4, f'O{row}')
    
    row += 1

# Summary at bottom
row += 2
ws4[f'A{row}'] = "END OF 2026 PROJECTIONS"
ws4[f'A{row}'].font = Font(bold=True, size=12)
ws4.merge_cells(f'A{row}:N{row}')

row += 1
ws4[f'B{row}'] = "Projected Total Debt (Dec 2026)"
ws4[f'N{row}'] = "=N16"
format_currency(ws4, f'N{row}')
ws4[f'N{row}'].font = Font(bold=True)

row += 1
ws4[f'B{row}'] = "Debt Paid Off"
ws4[f'N{row}'] = "=82108.87-N16"
format_currency(ws4, f'N{row}')
ws4[f'N{row}'].font = Font(bold=True, color="008000")

row += 1
ws4[f'B{row}'] = "Emergency Fund"
ws4[f'O{row}'] = "=O16"
format_currency(ws4, f'O{row}')
ws4[f'O{row}'].fill = SUCCESS_FILL


# ============== SHEET 5: ACTION PLAN ==============
ws5 = wb.create_sheet("Action Plan")
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 70

ws5['A1'] = "YOUR ACTION PLAN - GETTING OUT OF THIS MESS"
ws5['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws5.merge_cells('A1:B1')

actions = [
    ("WEEK 1", "EMERGENCY TRIAGE (Dec 27 - Jan 3)"),
    ("", "• Take $250 and put aside $100 for absolute emergencies"),
    ("", "• Call BMW Financial Services - you are $1,204 past due"),
    ("", "  - Ask about forbearance or payment arrangement"),
    ("", "  - Goal: Get them to stop any repo proceedings"),
    ("", "• Call LendKey - you are $111 past due"),
    ("", "  - Make minimum payment if you can"),
    ("", "• DO NOT use any more credit cards or BNPL"),
    ("", ""),
    ("JAN 2026", "STABILIZE"),
    ("", "• Your BNPL is ~$1,800 this month - that's painful but temporary"),
    ("", "• Pay all minimums on time - no more missed payments"),
    ("", "• Every dollar from paycheck goes to: Bills → Minimums → Food/Gas → Extra to BMW"),
    ("", "• Nelnet deferment ends in Feb - prepare for $226/mo increase"),
    ("", ""),
    ("FEB-APR", "DEBT AVALANCHE BEGINS"),
    ("", "• BNPL will drop significantly as loans close out"),
    ("", "• Once BMW is current, attack Best Buy (29.99% APR)"),
    ("", "• Target: Pay off Navy Federal $487 balance (quick win)"),
    ("", ""),
    ("MAY-AUG", "CREDIT CARD DESTRUCTION"),
    ("", "• BNPL should be mostly gone by now"),
    ("", "• All freed-up cash goes to Best Buy → Bank of America → Amex"),
    ("", "• Start building $1,000 emergency fund (20% of surplus)"),
    ("", ""),
    ("SEP-DEC", "MOMENTUM"),
    ("", "• Credit cards should be significantly lower"),
    ("", "• Emergency fund at $1,000"),
    ("", "• Continue avalanche method"),
    ("", ""),
    ("END 2026 GOALS", ""),
    ("", "✓ Emergency fund: $1,000"),
    ("", "✓ All accounts CURRENT (no past due)"),
    ("", "✓ BNPL: GONE"),
    ("", "✓ Credit cards: Reduced by 50%+ (~$8,000 remaining)"),
    ("", "✓ Total debt: Under $65,000"),
    ("", ""),
    ("RULES", "NON-NEGOTIABLE"),
    ("", "1. NO NEW DEBT - Cut up the credit cards if you have to"),
    ("", "2. Every dollar has a job BEFORE you spend it"),
    ("", "3. Check OpenFi every payday - know what's due"),
    ("", "4. $0 impulse purchases - if it's not budgeted, it doesn't happen"),
    ("", "5. If you get a windfall (tax refund, bonus), 100% to debt"),
]

row = 3
for period, action in actions:
    ws5[f'A{row}'] = period
    ws5[f'A{row}'].font = Font(bold=True) if period else BLACK_FONT
    ws5[f'B{row}'] = action
    if period in ["WEEK 1", "RULES"]:
        ws5[f'A{row}'].fill = ALERT_FILL
        ws5[f'B{row}'].fill = ALERT_FILL
    elif period in ["END 2026 GOALS"]:
        ws5[f'A{row}'].fill = SUCCESS_FILL
        ws5[f'B{row}'].fill = SUCCESS_FILL
    row += 1


# Save the workbook
output_path = "/Users/chris/projects/dev/openfinance/chris_financial_model_2026.xlsx"
wb.save(output_path)
print(f"Financial model saved to: {output_path}")
